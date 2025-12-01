import pandas as pd
import logging
from typing import List, Dict
from openpyxl import load_workbook
import io

logger = logging.getLogger(__name__)

def export_to_template_excel(data: List[Dict], template_path: str = "csv-templates/Data Sraping Template .xlsx") -> bytes:
    """
    Export scraped data to Excel using the provided template.
    Maps data to template columns and writes to appropriate sheets (Speakers, Exhibitors, Sponsors).
    """
    try:
        # Load the template
        template_wb = load_workbook(template_path)
        
        # Map our data structure to template columns
        template_columns = {
            'Contact First Name': 'first_name',
            'Contact Last Name': 'last_name',
            'Contact Position': 'job_title',
            'Contact Email  Address': 'email',
            'Contact Phone Number': 'phone',
            'Contact Linkedin URL ': 'linkedin_url',
            'Contact Country': 'country',
            'Organization Name': 'company_name',
            'Organization  Website Domain': 'company_website',
            'Organization Linkedin URL': '',  # We don't have this
            'Organization City': '',  # We don't have this
            'Organization  Country': 'country'  # Use contact country as fallback
        }
        
        # Convert data to DataFrame
        df = pd.DataFrame(data)
        
        # Create mapped data with template column names
        mapped_data = []
        for idx, row in df.iterrows():
            mapped_row = {}
            for template_col, our_col in template_columns.items():
                if our_col and our_col in row:
                    value = row[our_col]
                    # Handle NaN/None values
                    if pd.isna(value):
                        mapped_row[template_col] = ""
                    else:
                        mapped_row[template_col] = str(value) if value else ""
                else:
                    mapped_row[template_col] = ""
            mapped_data.append(mapped_row)
        
        # Categorize data by type
        speakers_data = []
        exhibitors_data = []
        sponsors_data = []
        decision_makers_data = []
        
        for idx, row in df.iterrows():
            category = str(row.get('category', '')).lower()
            mapped_row = mapped_data[idx]
            
            if 'decision maker' in category:
                decision_makers_data.append(mapped_row)
            elif 'speaker' in category:
                speakers_data.append(mapped_row)
            elif 'exhibitor' in category:
                exhibitors_data.append(mapped_row)
            elif 'sponsor' in category:
                sponsors_data.append(mapped_row)
            else:
                # Default to Speakers if unknown
                speakers_data.append(mapped_row)
        
        # Write to appropriate sheets
        sheets_to_update = {
            'Speakers': speakers_data + decision_makers_data,  # Include decision makers in Speakers
            'Exhibitors ': exhibitors_data,
            'Sponsors ': sponsors_data
        }
        
        for sheet_name, sheet_data in sheets_to_update.items():
            if sheet_name in template_wb.sheetnames:
                ws = template_wb[sheet_name]
                # Clear existing data (keep header row)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
                
                # Write new data starting from row 2 (row 1 is header)
                if sheet_data:
                    sheet_df = pd.DataFrame(sheet_data)
                    for idx, row in sheet_df.iterrows():
                        for col_idx, col_name in enumerate(template_columns.keys(), start=1):
                            cell_value = row.get(col_name, "")
                            ws.cell(row=idx+2, column=col_idx, value=cell_value)
        
        # Save to bytes buffer
        buffer = io.BytesIO()
        template_wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Exported {len(data)} records to template Excel file")
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error exporting to template: {e}")
        raise

